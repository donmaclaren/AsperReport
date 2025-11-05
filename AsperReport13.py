import streamlit as st
import re
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from io import TextIOWrapper
from zipfile import ZipFile, BadZipFile, is_zipfile
from openpyxl.styles import Alignment
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import DateAxis


def parse_temperature_line(line):
    if len(line.strip()) < 58:
        return None
    try:
        serial_num = int(line[0:4], 16)
        transmitter_no = int(line[4:6], 16)
        status_byte = int(line[6:8], 16)
        sign1 = -1 if line[8] == '-' else 1
        temp1 = sign1 * int(line[9:13]) / 10.0
        sign1u = -1 if line[13] == '-' else 1
        temp1u = sign1u * int(line[14:18]) / 10.0
        sign1l = -1 if line[18] == '-' else 1
        temp1l = sign1l * int(line[19:23]) / 10.0
        temp1d = int(line[23:27])
        sign2 = -1 if line[27] == '-' else 1
        temp2 = sign1 * int(line[28:32]) / 10.0
        sign2u = -1 if line[32] == '-' else 1
        temp2u = sign1u * int(line[33:37]) / 10.0
        sign2l = -1 if line[37] == '-' else 1
        temp2l = sign1l * int(line[38:42]) / 10.0
        temp2d = int(line[42:46])

        dt = line[46:58]
        dt_obj = datetime.strptime(dt, "%y%m%d%H%M%S")
        return {
            "Serial Num": serial_num,
            "Transmitter": transmitter_no,
            "Status": status_byte,
            "Temp (°C)": temp1,
            "Probe1 Max": temp1u,
            "Probe1 Min": temp1l,
            "Probe1 Del": temp1d,
            "Temp2 (°C)": temp2,
            "Probe2 Max": temp2u,
            "Probe2 Min": temp2l,
            "Probe2 Del": temp2d,
            "Date/Time": dt_obj
        }
    except Exception:
        return None


def extract_data_from_text(content):
    """Extract and parse the <Data> section from a single text file content."""
    match = re.search(r"<Data>(.*?)</Data>", content, re.DOTALL)
    if not match:
        return []
    data_section = match.group(1)
    lines = [l.strip() for l in data_section.splitlines() if l.strip() and not l.strip().startswith("$$")]
    parsed = [parse_temperature_line(l) for l in lines]
    return [p for p in parsed if p]

def extract_data_from_file(filename):
    def clean_text(s):
        return ''.join(c for c in s if 32 <= ord(c) <= 126 or c in '\r\n\t')

    with open(filename, "r", encoding="utf-8") as f:
        content = clean_text(f.read())
    match = re.search(r"<Data>(.*?)</Data>", content, re.DOTALL)
    if not match:
        raise ValueError("No <Data> section found.")
    data_section = match.group(1)
    lines = [l.strip() for l in data_section.splitlines() if l.strip() and not l.strip().startswith("$$")]
    parsed = [parse_temperature_line(l) for l in lines]
    return [p for p in parsed if p]


def extract_and_merge_from_zip(zip_filename):
    """Extract and merge all temperature data from text files inside a ZIP archive."""
    all_records = []
    with ZipFile(BytesIO(zip_bytes)) as zip_ref:
        for file_name in zip_ref.namelist():
            if file_name.lower().endswith(".txt"):
                with zip_ref.open(file_name) as file:
                    content = TextIOWrapper(file, encoding="utf-8", errors="ignore").read()
                    file_records = extract_data_from_text(content)
                    #for r in file_records:
                    #    r["SourceFile"] = file_name  # track where data came from
                    all_records.extend(file_records)
    return all_records







def create_excel_with_charts(data):
    if not data:
        print("⚠️ No valid temperature data found.")
        return


    df = pd.DataFrame(data)
    df = df.sort_values(["Transmitter", "Date/Time"],ascending=[True, True])

    grouped = df.groupby("Transmitter")

    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    
    for transmitter, group in grouped:

        #if transmitter == 5:
        #    break

        def safe_name(name):
            return ''.join(c for c in name if c.isalnum() or c in ('_', '-'))[:31]


        ws = wb.create_sheet(safe_name(f"TX_{transmitter}"))
        #ws.append(headers)
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=10)
        ws.cell(column=2, row=1, value=(f"Transmitter Report {group["Date/Time"].loc[group.index[0]]} to {group["Date/Time"].loc[group.index[-1]]}"))
        currentCell = ws['B1'] #or currentCell = ws['A1']
        currentCell.alignment = Alignment(horizontal='center')

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
        ws.cell(column=1, row=3, value=(f"TX no {transmitter}    Serial No {group["Serial Num"].loc[group.index[0]]}"))
        currentCell = ws['A3'] #or currentCell = ws['A1']
        currentCell.alignment = Alignment(horizontal='left')

        ws.merge_cells(start_row=5, start_column=1, end_row=9, end_column=5)
        ws.cell(column=1, row=5, value=(f"Probe 1 \nMax temp {group["Temp (°C)"].max()}\nMin temp {group["Temp (°C)"].min()}\nAverage temp {group["Temp (°C)"].mean().round(2)}"))
        currentCell = ws['A5'] #or currentCell = ws['A1']
        currentCell.alignment = Alignment(horizontal='left', vertical= 'top', wrapText=True)

        Prob2Disabled = False
        if group["Temp2 (°C)"].mean().round(2) == -999.9 or group["Temp2 (°C)"].mean().round(2) == 999.9:
            Prob2Disabled = True

        ws.merge_cells(start_row=5, start_column=6, end_row=9, end_column=10)
        if Prob2Disabled:
            ws.cell(column=6, row=5, value=(f"Probe 2 Disabled"))
        else:
            ws.cell(column=6, row=5, value=(f"Probe 2 \nMax temp {group["Temp2 (°C)"].max()}\nMin temp {group["Temp2 (°C)"].min()}\nAverage temp {group["Temp2 (°C)"].mean().round(2)}"))
 
        currentCell = ws['F5'] #or currentCell = ws['A1']
        currentCell.alignment = Alignment(horizontal='left', vertical= 'top', wrapText=True)

        ws.cell(column=1, row=10, value = "Data")

        # Write headers
        ws.append((list(group.columns))[2:])
        
        # Write rows
        for row in group.itertuples(index=False):
            ws.append((list(row))[2:])


        # Build chart
        max_row = ws.max_row
        
        chart = ScatterChart()
        chart.title = "Probe 1 Temperature"
        xvalues = Reference(ws, min_col=10, min_row=3, max_row=max_row)
        yvalues = Reference(ws, min_col=2, min_row=12, max_row=max_row)
        series = Series(values=yvalues, xvalues=xvalues, title="Probe 1 Temp")
        chart.series.append(series)
        chart.height = 10
        chart.y_axis.title = 'Temp'
        chart.x_axis.title = 'Time'
        chart.legend = None
        chart.x_axis.delete = True
        chart.y_axis.delete = False

        ws.add_chart(chart, f"A{max_row + 3}")


        # Build chart
        if Prob2Disabled == False:        
            chart2 = ScatterChart()
            chart2.title = "Probe 2 Temperature"
            xvalues = Reference(ws, min_col=10, min_row=3, max_row=max_row)
            yvalues = Reference(ws, min_col=6, min_row=12, max_row=max_row)
            series = Series(values=yvalues, xvalues=xvalues, title="Probe 2 Temp")
            chart2.series.append(series)
            chart2.height = 10
            chart2.y_axis.title = 'Temp'
            chart2.x_axis.title = 'Time'
            chart2.legend = None
            chart2.x_axis.delete = True
            chart2.y_axis.delete = False

            ws.add_chart(chart2, f"A{max_row + 23}")


    # Save workbook to memory
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio        

    

    #wb.save(output_filename)
    #print(f"✅ Excel file created successfully: {output_filename}")


# --- Streamlit UI ---
st.title("📈 Asper Temperature Analysis")
st.write("Upload a ZIP file containing temperature text files to merge them into one Excel report.")

uploaded_zip = st.file_uploader("Upload ZIP file", type=["zip"])

if uploaded_zip:
    try:
        zip_bytes = uploaded_zip.read()
        records = extract_and_merge_from_zip(zip_bytes)
        if not records:
            st.error("No valid temperature data found in ZIP.")
        else:
            #st.success(f"Parsed {len(records)} readings from {len(set([r['SourceFile'] for r in records]))} files.")
            st.success(f"Parsed {len(records)} readings.")

            excel_bytes = create_excel_with_charts(records)
            st.download_button(
                label="⬇️ Download Merged Excel File",
                data=excel_bytes,
                file_name="merged_temperature_data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # Optional: show sample preview
            st.dataframe(pd.DataFrame(records).head(20))
    except BadZipFile:
        st.error("Uploaded file is not a valid ZIP archive.")
